"""Document import endpoints for Excel/CSV file processing with intelligent detection."""

import os
import csv
import json
import re
import uuid as uuid_lib
from io import StringIO
from flask import Blueprint, request, jsonify, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from werkzeug.utils import secure_filename
from datetime import datetime
from openpyxl import load_workbook

from ..extensions import db
from ..models import Project, Workspace, OrganizationMember, Card, Column, Board
from ..services.ai_service import get_ai_service

imports_bp = Blueprint("imports", __name__)

UPLOAD_FOLDER = os.environ.get("UPLOAD_FOLDER", "uploads")
ALLOWED_EXTENSIONS = {"xlsx", "xls", "csv"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# In-memory storage for import sessions (in production, use Redis or database)
import_sessions = {}


def allowed_file(filename):
    """Check if file extension is allowed."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def check_project_access(project_id, user_id):
    """Check if user has access to project."""
    project = Project.query.get(project_id)
    if not project:
        return None, None

    workspace = Workspace.query.get(project.workspace_id)
    membership = OrganizationMember.query.filter_by(
        organization_id=workspace.organization_id, user_id=user_id
    ).first()

    return project, membership


def detect_data_type(value, context_hint=None):
    """Detect the data type of a cell value with optional context hint."""
    if value is None or value == "":
        return "empty"

    value_str = str(value).strip()

    # Check for priority patterns
    if value_str.upper() in ["P0", "P1", "P2", "P3", "P4", "HIGH", "MEDIUM", "LOW", "CRITICAL"]:
        return "priority"

    # Check for status patterns
    status_keywords = {
        "TODO", "TO DO", "TO-DO", "DONE", "IN PROGRESS", "IN-PROGRESS",
        "COMPLETE", "COMPLETED", "PENDING", "BLOCKED", "OPEN", "CLOSED",
        "NOT STARTED", "IN REVIEW", "REVIEW", "QA", "TESTING"
    }
    if value_str.upper() in status_keywords:
        return "status"

    # Check for percentage (completion)
    if re.match(r"^\d{1,3}%$", value_str):
        return "percentage"

    # Check for time/duration patterns (2h, 30m, 2.5hrs, etc.)
    if re.match(r"^\d+\.?\d*\s*(h|hr|hrs|hours?|m|min|mins|minutes?|d|days?|w|weeks?|pts?|points?)$", value_str.lower()):
        return "duration"

    # Check for number
    try:
        num = float(value_str.replace(",", ""))
        # If context hint suggests this is a row number, mark it
        if context_hint == "row_identifier" and num == int(num) and 1 <= num <= 1000:
            return "row_identifier"
        return "number"
    except ValueError:
        pass

    # Check for date with multiple patterns
    date_patterns = [
        "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d",
        "%d-%m-%Y", "%m-%d-%Y", "%b %d, %Y", "%B %d, %Y",
        "%d %b %Y", "%d %B %Y", "%Y-%m-%d %H:%M:%S"
    ]
    for pattern in date_patterns:
        try:
            datetime.strptime(value_str, pattern)
            return "date"
        except ValueError:
            continue

    # Check for email
    if "@" in value_str and "." in value_str.split("@")[-1]:
        return "email"

    # Check for person name pattern (First Last, initials, etc.)
    if context_hint == "person" and re.match(r"^[A-Z][a-z]+ [A-Z][a-z]+$", value_str):
        return "person"

    # Default to text
    return "text"


def calculate_header_score(row, row_idx):
    """Calculate a score for how likely a row is to be a header row."""
    if not row or all(cell is None for cell in row):
        return -1

    score = 0
    non_empty_cells = [cell for cell in row if cell is not None and str(cell).strip()]

    if len(non_empty_cells) < 2:
        return -1

    # Header keywords that strongly indicate a header row
    header_keywords = {
        "task", "title", "name", "description", "status", "priority",
        "assignee", "owner", "responsible", "date", "due", "deadline",
        "estimate", "points", "hours", "effort", "start", "end",
        "id", "number", "type", "category", "label", "tag", "notes",
        "comment", "progress", "completion", "phase", "stage", "column"
    }

    for cell in non_empty_cells:
        cell_str = str(cell).lower().strip()

        # Check for header keywords
        for keyword in header_keywords:
            if keyword in cell_str:
                score += 20
                break

        # Headers are usually short
        if 2 <= len(cell_str) <= 30:
            score += 5

        # Headers usually don't start with numbers (unless it's like "1. Task")
        if not cell_str[0].isdigit():
            score += 3

        # Headers are usually unique text, not dates or numbers
        if detect_data_type(cell) == "text":
            score += 5
        elif detect_data_type(cell) in ["number", "date"]:
            score -= 10

    # Penalty for rows that are too far down (headers usually in first 10 rows)
    if row_idx > 10:
        score -= (row_idx - 10) * 5

    # Bonus for row having many columns filled
    fill_ratio = len(non_empty_cells) / len(row) if row else 0
    score += int(fill_ratio * 20)

    return score


def detect_header_row(raw_rows, max_check=15):
    """Intelligently detect which row contains the headers."""
    if not raw_rows:
        return 0, []

    best_row_idx = 0
    best_score = -1

    # Check first N rows to find the best header candidate
    for idx, row in enumerate(raw_rows[:max_check]):
        score = calculate_header_score(row, idx)
        if score > best_score:
            best_score = score
            best_row_idx = idx

    return best_row_idx, raw_rows[best_row_idx] if best_row_idx < len(raw_rows) else []


def detect_hierarchical_structure(rows, headers):
    """Detect if the data has a hierarchical structure (parent/child tasks)."""
    structure = {
        "is_hierarchical": False,
        "id_column": None,
        "parent_column": None,
        "level_indicator": None,  # indentation, numbering, etc.
        "hierarchy_pattern": None
    }

    if not rows or not headers:
        return structure

    # Look for ID columns (columns with sequential numbers 1, 2, 3...)
    for header in headers:
        values = [row.get(header) for row in rows[:20] if row.get(header) is not None]
        if not values:
            continue

        # Check for sequential integers
        try:
            int_values = [int(float(str(v))) for v in values if str(v).strip()]
            if len(int_values) >= 3:
                # Check if mostly sequential
                sequential_count = sum(1 for i in range(1, len(int_values))
                                       if int_values[i] == int_values[i-1] + 1)
                if sequential_count >= len(int_values) * 0.5:
                    structure["id_column"] = header
                    break
        except (ValueError, TypeError):
            pass

    # Look for parent ID columns
    parent_patterns = ["parent", "parent_id", "parent id", "parent task", "depends on", "subtask of"]
    for header in headers:
        header_lower = header.lower()
        for pattern in parent_patterns:
            if pattern in header_lower:
                structure["parent_column"] = header
                structure["is_hierarchical"] = True
                break

    # Check for indentation-based hierarchy (text starting with spaces/dashes)
    for header in headers:
        if "task" in header.lower() or "title" in header.lower() or "name" in header.lower():
            values = [str(row.get(header, "")) for row in rows[:20]]
            indented_count = sum(1 for v in values if v.startswith("  ") or v.startswith("-") or v.startswith("•"))
            if indented_count >= 3:
                structure["is_hierarchical"] = True
                structure["level_indicator"] = "indentation"
                break

    # Check for numbering hierarchy (1, 1.1, 1.2, 2, 2.1, etc.)
    for header in headers:
        values = [str(row.get(header, "")) for row in rows[:30]]
        # Pattern for hierarchical numbering
        numbered_count = sum(1 for v in values if re.match(r"^\d+(\.\d+)*\.?$", v.strip()))
        if numbered_count >= 5:
            structure["is_hierarchical"] = True
            structure["hierarchy_pattern"] = "numbered"
            if not structure["id_column"]:
                structure["id_column"] = header
            break

    return structure


def detect_adjacent_content_columns(headers, rows, data_types):
    """Detect when title content is split across adjacent columns."""
    adjacent_pairs = []

    for i, header in enumerate(headers[:-1]):
        next_header = headers[i + 1]

        # Check if first column has short content (numbers, IDs) and next has longer text
        col1_values = [str(row.get(header, "")) for row in rows[:15] if row.get(header)]
        col2_values = [str(row.get(next_header, "")) for row in rows[:15] if row.get(next_header)]

        if not col1_values or not col2_values:
            continue

        col1_avg_len = sum(len(v) for v in col1_values) / len(col1_values)
        col2_avg_len = sum(len(v) for v in col2_values) / len(col2_values)

        # Check if col1 is numeric/short and col2 is longer text
        col1_numeric = all(v.replace(".", "").replace("-", "").isdigit() for v in col1_values if v.strip())

        if col1_numeric and col1_avg_len < 10 and col2_avg_len > 20:
            # This looks like ID + Description pattern
            adjacent_pairs.append({
                "id_column": header,
                "content_column": next_header,
                "pattern": "id_description",
                "confidence": 0.9
            })
        elif col1_avg_len < 5 and col2_avg_len > 30:
            # Short prefix + long description
            adjacent_pairs.append({
                "id_column": header,
                "content_column": next_header,
                "pattern": "prefix_content",
                "confidence": 0.7
            })

    return adjacent_pairs


def smart_generate_title(row, mapping, adjacent_pairs):
    """Generate a smart title when the mapped title column has poor content."""
    title_col = mapping.get("title")
    title_value = row.get(title_col) if title_col else None

    if title_value:
        title_str = str(title_value).strip()
        # If title is just a number or very short, try to enhance it
        if title_str.isdigit() or len(title_str) <= 3:
            # Check if we have adjacent column content
            for pair in adjacent_pairs:
                if pair["id_column"] == title_col:
                    content = row.get(pair["content_column"])
                    if content:
                        return str(content).strip()[:200]

            # Try description as fallback
            desc_col = mapping.get("description")
            if desc_col:
                desc = row.get(desc_col)
                if desc:
                    desc_str = str(desc).strip()
                    # Use first sentence or first 100 chars as title
                    first_sentence = desc_str.split(".")[0]
                    if len(first_sentence) > 10:
                        return first_sentence[:200]

        return title_str[:200]

    return None


def analyze_column_patterns(headers, rows):
    """Deep analysis of column content patterns for better detection."""
    analysis = {}

    for header in headers:
        values = [row.get(header) for row in rows[:30] if row.get(header) is not None]
        if not values:
            analysis[header] = {"type": "empty", "confidence": 1.0}
            continue

        str_values = [str(v).strip() for v in values]
        non_empty = [v for v in str_values if v]

        if not non_empty:
            analysis[header] = {"type": "empty", "confidence": 1.0}
            continue

        # Analyze patterns
        avg_len = sum(len(v) for v in non_empty) / len(non_empty)
        unique_ratio = len(set(non_empty)) / len(non_empty)

        # Check for sequential numbers
        try:
            nums = [int(float(v)) for v in non_empty if v.replace(".", "").replace("-", "").isdigit()]
            if len(nums) >= 3 and len(nums) / len(non_empty) > 0.8:
                is_sequential = all(nums[i] == nums[i-1] + 1 for i in range(1, len(nums)))
                if is_sequential:
                    analysis[header] = {
                        "type": "sequential_id",
                        "confidence": 0.95,
                        "sample": nums[:5]
                    }
                    continue
        except (ValueError, TypeError):
            pass

        # Check for status values
        status_keywords = {"todo", "done", "in progress", "complete", "pending", "blocked", "open", "closed"}
        status_matches = sum(1 for v in non_empty if v.lower() in status_keywords)
        if status_matches / len(non_empty) > 0.5:
            analysis[header] = {"type": "status", "confidence": status_matches / len(non_empty)}
            continue

        # Check for names (person names)
        name_pattern = sum(1 for v in non_empty if re.match(r"^[A-Z][a-z]+ [A-Z][a-z]+$", v))
        if name_pattern / len(non_empty) > 0.5:
            analysis[header] = {"type": "person_name", "confidence": name_pattern / len(non_empty)}
            continue

        # Check for time estimates (2h, 16 pts, etc.)
        time_pattern = sum(1 for v in non_empty if re.match(r"^\d+\.?\d*\s*(h|hr|hrs|m|min|pts?|points?)$", v.lower()))
        if time_pattern / len(non_empty) > 0.4:
            analysis[header] = {"type": "time_estimate", "confidence": time_pattern / len(non_empty)}
            continue

        # Determine if this is likely a title column
        if 10 < avg_len < 150 and unique_ratio > 0.7:
            analysis[header] = {
                "type": "title_candidate",
                "confidence": unique_ratio * 0.8,
                "avg_length": avg_len
            }
        elif avg_len > 100:
            analysis[header] = {
                "type": "description_candidate",
                "confidence": 0.8,
                "avg_length": avg_len
            }
        else:
            analysis[header] = {
                "type": "text",
                "confidence": 0.5,
                "avg_length": avg_len,
                "unique_ratio": unique_ratio
            }

    return analysis


def parse_excel_raw(file_path):
    """Parse Excel file and return ALL rows including potential header rows."""
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active

    all_rows = []
    for row in sheet.iter_rows(values_only=True):
        all_rows.append(list(row))

    wb.close()
    return all_rows


def parse_csv_raw(file_path):
    """Parse CSV file and return ALL rows including potential header rows."""
    all_rows = []
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        # Try to detect dialect
        sample = f.read(4096)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(f, dialect=dialect)
        for row in reader:
            all_rows.append(row)

    return all_rows


def process_raw_to_structured(raw_rows, header_row_idx):
    """Convert raw rows to structured data with detected headers."""
    if header_row_idx >= len(raw_rows):
        return [], []

    header_row = raw_rows[header_row_idx]
    headers = [str(cell) if cell else f"Column {i+1}" for i, cell in enumerate(header_row)]

    rows = []
    for row in raw_rows[header_row_idx + 1:]:
        # Skip empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = cell
        rows.append(row_dict)

    return headers, rows


def smart_detect_column_mapping(headers, rows, data_types, column_analysis=None):
    """AI-like smart detection of column mappings with advanced pattern recognition."""
    mapping = {
        "title": None,
        "description": None,
        "priority": None,
        "story_points": None,
        "assignee": None,
        "due_date": None,
        "status": None,
        "labels": None,
        "start_date": None,
        "estimate": None,
    }

    headers_lower = {h.lower().strip(): h for h in headers}
    column_analysis = column_analysis or analyze_column_patterns(headers, rows)

    # ============ TITLE DETECTION ============
    # Method 1: Look for explicit title patterns
    title_patterns = ["title", "name", "task", "item", "issue", "ticket", "summary", "subject", "todo", "card", "work item"]
    for pattern in title_patterns:
        for h_lower, h_orig in headers_lower.items():
            if pattern in h_lower and "description" not in h_lower:
                mapping["title"] = h_orig
                break
        if mapping["title"]:
            break

    # Method 2: Use column analysis to find best title candidate
    if not mapping["title"]:
        best_title_score = 0
        for header, analysis in column_analysis.items():
            if analysis.get("type") == "title_candidate":
                score = analysis.get("confidence", 0)
                if score > best_title_score:
                    best_title_score = score
                    mapping["title"] = header

    # Method 3: Find first text column with good uniqueness
    if not mapping["title"]:
        for header in headers:
            if data_types.get(header) == "text":
                sample_values = [str(row.get(header, "")) for row in rows[:15] if row.get(header)]
                if not sample_values:
                    continue
                avg_len = sum(len(v) for v in sample_values) / len(sample_values)
                unique_ratio = len(set(sample_values)) / len(sample_values) if sample_values else 0
                if 5 < avg_len < 200 and unique_ratio > 0.7:
                    mapping["title"] = header
                    break

    # ============ DESCRIPTION DETECTION ============
    desc_patterns = ["description", "desc", "detail", "details", "body", "content", "notes", "note", "comment", "remarks"]
    for pattern in desc_patterns:
        for h_lower, h_orig in headers_lower.items():
            if pattern in h_lower and h_orig != mapping["title"]:
                mapping["description"] = h_orig
                break
        if mapping["description"]:
            break

    # Find longest text column as description
    if not mapping["description"]:
        max_avg_len = 0
        for header in headers:
            if header == mapping["title"]:
                continue
            analysis = column_analysis.get(header, {})
            if analysis.get("type") in ["description_candidate", "text"]:
                avg_len = analysis.get("avg_length", 0)
                if avg_len > max_avg_len and avg_len > 30:
                    max_avg_len = avg_len
                    mapping["description"] = header

    # ============ PRIORITY DETECTION ============
    prio_patterns = ["priority", "prio", "urgency", "importance", "severity", "p0", "p1", "p2", "level"]
    for pattern in prio_patterns:
        for h_lower, h_orig in headers_lower.items():
            if pattern in h_lower:
                mapping["priority"] = h_orig
                break
        if mapping["priority"]:
            break

    # Check content for priority values
    if not mapping["priority"]:
        for header in headers:
            sample_values = [str(row.get(header, "")).upper().strip() for row in rows[:15] if row.get(header)]
            priority_keywords = {"P0", "P1", "P2", "P3", "P4", "HIGH", "MEDIUM", "LOW", "CRITICAL", "URGENT", "NORMAL"}
            if sample_values and sum(1 for v in sample_values if v in priority_keywords) / len(sample_values) > 0.4:
                mapping["priority"] = header
                break

    # ============ STORY POINTS / ESTIMATE DETECTION ============
    points_patterns = ["point", "points", "story_point", "estimate", "effort", "size", "sp", "complexity", "hrs", "hours", "plan"]
    for pattern in points_patterns:
        for h_lower, h_orig in headers_lower.items():
            if pattern in h_lower:
                mapping["story_points"] = h_orig
                break
        if mapping["story_points"]:
            break

    # Check column analysis for time estimates
    if not mapping["story_points"]:
        for header, analysis in column_analysis.items():
            if analysis.get("type") == "time_estimate":
                mapping["story_points"] = header
                mapping["estimate"] = header
                break

    # Check for fibonacci-like numbers
    if not mapping["story_points"]:
        fibonacci_like = {1, 2, 3, 5, 8, 13, 21}
        for header in headers:
            if data_types.get(header) == "number":
                sample_values = []
                for row in rows[:15]:
                    try:
                        val = int(float(str(row.get(header, 0))))
                        sample_values.append(val)
                    except:
                        pass
                if sample_values and all(v in fibonacci_like or 0 <= v <= 21 for v in sample_values):
                    mapping["story_points"] = header
                    break

    # ============ ASSIGNEE DETECTION ============
    assignee_patterns = ["assignee", "assigned", "owner", "responsible", "user", "person", "member", "resource"]
    for pattern in assignee_patterns:
        for h_lower, h_orig in headers_lower.items():
            if pattern in h_lower:
                mapping["assignee"] = h_orig
                break
        if mapping["assignee"]:
            break

    # Check for person name columns
    if not mapping["assignee"]:
        for header, analysis in column_analysis.items():
            if analysis.get("type") == "person_name":
                mapping["assignee"] = header
                break

    # ============ STATUS DETECTION ============
    status_patterns = ["status", "state", "stage", "column", "progress", "phase"]
    for pattern in status_patterns:
        for h_lower, h_orig in headers_lower.items():
            if pattern in h_lower:
                mapping["status"] = h_orig
                break
        if mapping["status"]:
            break

    # Check column analysis for status
    if not mapping["status"]:
        for header, analysis in column_analysis.items():
            if analysis.get("type") == "status":
                mapping["status"] = header
                break

    # ============ DATE DETECTION ============
    start_patterns = ["start", "begin", "created"]
    end_patterns = ["due", "deadline", "end", "target", "finish", "complete by"]

    for pattern in end_patterns:
        for h_lower, h_orig in headers_lower.items():
            if pattern in h_lower:
                mapping["due_date"] = h_orig
                break
        if mapping["due_date"]:
            break

    for pattern in start_patterns:
        for h_lower, h_orig in headers_lower.items():
            if pattern in h_lower and h_orig != mapping["due_date"]:
                mapping["start_date"] = h_orig
                break
        if mapping["start_date"]:
            break

    # ============ LABELS DETECTION ============
    label_patterns = ["label", "tag", "category", "type", "kind", "group", "epic", "theme"]
    for pattern in label_patterns:
        for h_lower, h_orig in headers_lower.items():
            if pattern in h_lower:
                mapping["labels"] = h_orig
                break
        if mapping["labels"]:
            break

    # ============ CONFIDENCE CALCULATION ============
    confidence = 0
    confidence_details = {}

    if mapping["title"]:
        confidence += 40
        confidence_details["title"] = "found"
    if mapping["description"]:
        confidence += 15
        confidence_details["description"] = "found"
    if mapping["priority"]:
        confidence += 10
        confidence_details["priority"] = "found"
    if mapping["story_points"]:
        confidence += 10
        confidence_details["story_points"] = "found"
    if mapping["status"]:
        confidence += 10
        confidence_details["status"] = "found"
    if mapping["due_date"]:
        confidence += 5
        confidence_details["due_date"] = "found"
    if mapping["assignee"]:
        confidence += 5
        confidence_details["assignee"] = "found"
    if mapping["start_date"]:
        confidence += 3
        confidence_details["start_date"] = "found"
    if mapping["labels"]:
        confidence += 2
        confidence_details["labels"] = "found"

    return mapping, confidence, confidence_details


@imports_bp.route("/upload", methods=["POST"])
@jwt_required()
def upload_file():
    """Upload and parse an Excel or CSV file with intelligent detection."""
    user_id = get_jwt_identity()
    project_id = request.form.get("project_id")

    if not project_id:
        return jsonify({"error": "project_id required"}), 400

    project, membership = check_project_access(project_id, user_id)
    if not membership:
        return jsonify({"error": "Forbidden"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]

    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": "File type not allowed. Supported: xlsx, xls, csv"}), 400

    # Check file size
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)

    if file_size > MAX_FILE_SIZE:
        return jsonify({"error": "File too large. Maximum size is 10MB"}), 400

    # Generate unique import ID and save file
    import_id = str(uuid_lib.uuid4())
    ext = file.filename.rsplit(".", 1)[1].lower()
    secure_name = f"{import_id}.{ext}"

    # Create uploads directory if needed
    upload_dir = os.path.join(UPLOAD_FOLDER, "imports")
    os.makedirs(upload_dir, exist_ok=True)

    file_path = os.path.join(upload_dir, secure_name)
    file.save(file_path)

    try:
        # Parse raw data first (all rows)
        if ext in ["xlsx", "xls"]:
            raw_rows = parse_excel_raw(file_path)
        else:
            raw_rows = parse_csv_raw(file_path)

        if not raw_rows or len(raw_rows) == 0:
            os.remove(file_path)
            return jsonify({"error": "File appears to be empty"}), 400

        # Intelligent header row detection
        header_row_idx, detected_header = detect_header_row(raw_rows)

        # Process into structured data
        headers, rows = process_raw_to_structured(raw_rows, header_row_idx)

        if not headers or len(headers) == 0:
            os.remove(file_path)
            return jsonify({"error": "Could not detect headers in file"}), 400

        # Deep column analysis
        column_analysis = analyze_column_patterns(headers, rows)

        # Detect hierarchical structure
        hierarchy_info = detect_hierarchical_structure(rows, headers)

        # Detect adjacent content columns (ID + Description pairs)
        adjacent_pairs = detect_adjacent_content_columns(headers, rows, {})

        # Detect data types
        data_types = {}
        for header in headers:
            sample_values = [row.get(header) for row in rows[:15] if row.get(header)]
            types = [detect_data_type(v) for v in sample_values]
            non_empty_types = [t for t in types if t != "empty"]
            if non_empty_types:
                data_types[header] = max(set(non_empty_types), key=non_empty_types.count)
            else:
                data_types[header] = "text"

        structure = {
            "headers": headers,
            "data_types": data_types,
            "row_count": len(rows),
            "column_count": len(headers),
            "header_row_detected": header_row_idx + 1,  # 1-indexed for user display
            "column_analysis": column_analysis,
            "hierarchy_detected": hierarchy_info["is_hierarchical"],
            "hierarchy_info": hierarchy_info,
            "adjacent_content_pairs": adjacent_pairs,
        }

        # Smart detect column mapping with all intelligence
        smart_mapping, confidence, confidence_details = smart_detect_column_mapping(
            headers, rows, data_types, column_analysis
        )

        # Store session data
        import_sessions[import_id] = {
            "id": import_id,
            "project_id": project_id,
            "user_id": user_id,
            "file_path": file_path,
            "file_type": ext,
            "original_filename": file.filename,
            "file_size": file_size,
            "raw_rows": raw_rows,
            "header_row_idx": header_row_idx,
            "headers": headers,
            "rows": rows,
            "structure": structure,
            "smart_mapping": smart_mapping,
            "mapping_confidence": confidence,
            "confidence_details": confidence_details,
            "column_analysis": column_analysis,
            "hierarchy_info": hierarchy_info,
            "adjacent_pairs": adjacent_pairs,
            "status": "preview",
            "created_at": datetime.now().isoformat(),
        }

        # Return preview data with all intelligent detection info
        return jsonify({
            "import_id": import_id,
            "filename": file.filename,
            "file_type": ext,
            "structure": structure,
            "preview_rows": rows[:15],  # First 15 rows for preview
            "smart_mapping": smart_mapping,
            "mapping_confidence": confidence,
            "confidence_details": confidence_details,
            "insights": {
                "header_row": header_row_idx + 1,
                "is_hierarchical": hierarchy_info["is_hierarchical"],
                "hierarchy_pattern": hierarchy_info.get("hierarchy_pattern"),
                "adjacent_content_detected": len(adjacent_pairs) > 0,
                "adjacent_pairs": adjacent_pairs,
            }
        }), 201

    except Exception as e:
        # Cleanup on error
        if os.path.exists(file_path):
            os.remove(file_path)
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 400


@imports_bp.route("/<import_id>/preview", methods=["GET"])
@jwt_required()
def get_preview(import_id):
    """Get preview data for an import session."""
    user_id = get_jwt_identity()

    session = import_sessions.get(import_id)
    if not session:
        return jsonify({"error": "Import session not found"}), 404

    if session["user_id"] != user_id:
        return jsonify({"error": "Forbidden"}), 403

    return jsonify({
        "import_id": import_id,
        "filename": session["original_filename"],
        "file_type": session["file_type"],
        "structure": session["structure"],
        "preview_rows": session["rows"][:20],
    })


@imports_bp.route("/<import_id>/process", methods=["POST"])
@jwt_required()
def process_import(import_id):
    """Process import with intelligent task extraction and AI enhancement."""
    user_id = get_jwt_identity()

    session = import_sessions.get(import_id)
    if not session:
        return jsonify({"error": "Import session not found"}), 404

    if session["user_id"] != user_id:
        return jsonify({"error": "Forbidden"}), 403

    # Get column mapping from request or use smart detection
    data = request.json or {}
    column_mapping = data.get("column_mapping") or session.get("smart_mapping", {})
    use_ai = data.get("use_ai", True)
    use_smart_titles = data.get("use_smart_titles", True)

    # Get adjacent pairs for smart title generation
    adjacent_pairs = session.get("adjacent_pairs", [])
    hierarchy_info = session.get("hierarchy_info", {})

    # Extract tasks from rows with intelligent processing
    tasks = []
    parent_task_map = {}  # For hierarchical task reconstruction

    for row_idx, row in enumerate(session["rows"]):
        # Smart title extraction
        title = None
        title_col = column_mapping.get("title")

        if use_smart_titles:
            title = smart_generate_title(row, column_mapping, adjacent_pairs)
        else:
            title = row.get(title_col) if title_col else None

        # Skip rows without meaningful content
        if not title or str(title).strip() == "":
            # Try to get title from adjacent content column
            for pair in adjacent_pairs:
                content = row.get(pair.get("content_column"))
                if content and str(content).strip():
                    title = str(content).strip()
                    break

        if not title or str(title).strip() == "":
            continue

        title_str = str(title).strip()

        # Skip if title is just a number and we have no other content
        if title_str.isdigit() and len(title_str) <= 3:
            # Check if we have actual content elsewhere
            has_content = False
            for header in session["headers"]:
                val = row.get(header)
                if val and str(val).strip() and not str(val).strip().isdigit():
                    has_content = True
                    break
            if not has_content:
                continue

        task = {
            "title": title_str[:500],
            "description": "",
            "priority": None,
            "story_points": None,
            "status": None,
            "assignee": None,
            "due_date": None,
            "start_date": None,
            "labels": [],
            "parent_id": None,
            "hierarchy_level": 0,
            "original_row": row_idx + 1,
        }

        # Map description
        if column_mapping.get("description"):
            desc = row.get(column_mapping["description"])
            if desc:
                task["description"] = str(desc).strip()[:5000]

        # Map priority with smart conversion
        if column_mapping.get("priority"):
            prio = row.get(column_mapping["priority"])
            if prio:
                prio_str = str(prio).upper().strip()
                if prio_str in ["P0", "P1", "P2", "P3", "P4"]:
                    task["priority"] = prio_str
                elif prio_str in ["CRITICAL", "URGENT", "HIGHEST"]:
                    task["priority"] = "P0"
                elif prio_str == "HIGH":
                    task["priority"] = "P1"
                elif prio_str in ["MEDIUM", "NORMAL", "MED"]:
                    task["priority"] = "P2"
                elif prio_str == "LOW":
                    task["priority"] = "P3"
                elif prio_str in ["LOWEST", "MINOR"]:
                    task["priority"] = "P4"

        # Map story points / estimate with smart parsing
        if column_mapping.get("story_points"):
            points = row.get(column_mapping["story_points"])
            if points:
                points_str = str(points).lower().strip()
                # Parse various formats: "16 pts", "2h", "3 hours", "5", etc.
                match = re.match(r"^(\d+\.?\d*)\s*(pts?|points?|h|hr|hrs|hours?)?", points_str)
                if match:
                    try:
                        task["story_points"] = int(float(match.group(1)))
                    except (ValueError, TypeError):
                        pass

        # Map status
        if column_mapping.get("status"):
            status = row.get(column_mapping["status"])
            if status:
                task["status"] = str(status).strip()

        # Map assignee
        if column_mapping.get("assignee"):
            assignee = row.get(column_mapping["assignee"])
            if assignee:
                task["assignee"] = str(assignee).strip()

        # Map dates
        if column_mapping.get("due_date"):
            due = row.get(column_mapping["due_date"])
            if due:
                task["due_date"] = str(due).strip()

        if column_mapping.get("start_date"):
            start = row.get(column_mapping["start_date"])
            if start:
                task["start_date"] = str(start).strip()

        # Map labels (handle comma-separated)
        if column_mapping.get("labels"):
            labels = row.get(column_mapping["labels"])
            if labels:
                labels_str = str(labels).strip()
                if "," in labels_str:
                    task["labels"] = [l.strip() for l in labels_str.split(",") if l.strip()]
                else:
                    task["labels"] = [labels_str] if labels_str else []

        # Handle hierarchical structure
        if hierarchy_info.get("is_hierarchical"):
            id_col = hierarchy_info.get("id_column")
            if id_col:
                task_id = row.get(id_col)
                if task_id:
                    task_id_str = str(task_id).strip()
                    # Check for hierarchical numbering (1.1, 1.2, 2.1, etc.)
                    if "." in task_id_str:
                        parts = task_id_str.split(".")
                        task["hierarchy_level"] = len(parts) - 1
                        if len(parts) > 1:
                            parent_id = ".".join(parts[:-1])
                            task["parent_id"] = parent_id
                    parent_task_map[task_id_str] = len(tasks)

        tasks.append(task)

    # AI enhancement if enabled
    ai_suggestions = None
    ai_analysis = None
    if use_ai and tasks:
        ai_service = get_ai_service()
        if ai_service.is_enabled():
            # Send task summary for AI analysis
            ai_suggestions = ai_service.enhance_imported_tasks(tasks[:25])

            # Also try to get AI analysis of the overall structure
            ai_analysis = ai_analyze_import_structure(tasks, session.get("structure", {}))

    # Update session
    session["extracted_tasks"] = tasks
    session["column_mapping"] = column_mapping
    session["ai_suggestions"] = ai_suggestions
    session["ai_analysis"] = ai_analysis
    session["status"] = "processed"

    return jsonify({
        "import_id": import_id,
        "task_count": len(tasks),
        "tasks": tasks[:75],  # Return first 75 for preview
        "ai_suggestions": ai_suggestions,
        "ai_analysis": ai_analysis,
        "column_mapping": column_mapping,
        "hierarchy_detected": hierarchy_info.get("is_hierarchical", False),
        "processing_insights": {
            "tasks_extracted": len(tasks),
            "tasks_with_description": sum(1 for t in tasks if t.get("description")),
            "tasks_with_priority": sum(1 for t in tasks if t.get("priority")),
            "tasks_with_points": sum(1 for t in tasks if t.get("story_points")),
            "tasks_with_assignee": sum(1 for t in tasks if t.get("assignee")),
            "hierarchical_tasks": sum(1 for t in tasks if t.get("hierarchy_level", 0) > 0),
        }
    })


def ai_analyze_import_structure(tasks, structure):
    """Use AI to analyze the import structure and provide insights."""
    ai_service = get_ai_service()
    if not ai_service.is_enabled():
        return None

    try:
        # Prepare summary for AI
        task_summary = []
        for t in tasks[:20]:
            task_summary.append({
                "title": t.get("title", "")[:100],
                "has_description": bool(t.get("description")),
                "priority": t.get("priority"),
                "points": t.get("story_points"),
                "status": t.get("status"),
            })

        prompt = f"""Analyze this imported task list and provide insights:

IMPORT STATISTICS:
- Total tasks: {len(tasks)}
- Column count: {structure.get('column_count', 'unknown')}
- Header row: {structure.get('header_row_detected', 'row 1')}

SAMPLE TASKS:
{json.dumps(task_summary, indent=2)}

Provide analysis in JSON format:
{{
    "data_quality_score": <1-10>,
    "completeness": {{
        "has_good_titles": true/false,
        "has_descriptions": true/false,
        "has_estimates": true/false,
        "has_priorities": true/false
    }},
    "recommendations": [
        "recommendation 1",
        "recommendation 2"
    ],
    "detected_project_type": "software|marketing|design|operations|general",
    "suggested_workflow": "kanban|scrum|simple"
}}"""

        from openai import OpenAI
        client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "You are a project management expert analyzing imported task data. Provide concise, actionable insights."
                },
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            max_tokens=500,
        )

        return json.loads(response.choices[0].message.content)
    except Exception as e:
        import logging
        logging.error(f"AI import analysis failed: {e}")
        return None


@imports_bp.route("/<import_id>/confirm", methods=["POST"])
@jwt_required()
def confirm_import(import_id):
    """Create cards from confirmed import tasks."""
    user_id = get_jwt_identity()

    session = import_sessions.get(import_id)
    if not session:
        return jsonify({"error": "Import session not found"}), 404

    if session["user_id"] != user_id:
        return jsonify({"error": "Forbidden"}), 403

    data = request.json or {}
    approved_tasks = data.get("tasks", session.get("extracted_tasks", []))
    board_id = data.get("board_id")
    column_id = data.get("column_id")

    if not board_id:
        # Get the first board in the project
        project_id = session["project_id"]
        board = Board.query.filter_by(project_id=project_id).first()
        if not board:
            return jsonify({"error": "No board found in project. Create a board first."}), 400
        board_id = board.id

    if not column_id:
        # Get the first column (usually "Backlog" or "To Do")
        column = Column.query.filter_by(board_id=board_id).order_by(Column.position).first()
        if not column:
            return jsonify({"error": "No column found in board. Create a column first."}), 400
        column_id = column.id

    # Create cards
    created_cards = []
    max_position = db.session.query(db.func.max(Card.position)).filter_by(column_id=column_id).scalar() or 0

    for i, task in enumerate(approved_tasks):
        if not task.get("title"):
            continue

        card = Card(
            column_id=column_id,
            title=task["title"][:500],  # Limit title length
            description=task.get("description", "")[:5000] if task.get("description") else None,
            priority=task.get("priority"),
            story_points=task.get("story_points"),
            position=max_position + i + 1,
            created_by=user_id,
        )
        db.session.add(card)
        created_cards.append(card)

    db.session.commit()

    # Cleanup session
    if os.path.exists(session["file_path"]):
        os.remove(session["file_path"])
    del import_sessions[import_id]

    return jsonify({
        "message": f"Successfully created {len(created_cards)} cards",
        "created_count": len(created_cards),
        "card_ids": [str(c.id) for c in created_cards],
        "board_id": str(board_id),
        "column_id": str(column_id),
    })


@imports_bp.route("/<import_id>", methods=["DELETE"])
@jwt_required()
def cancel_import(import_id):
    """Cancel an import session and cleanup files."""
    user_id = get_jwt_identity()

    session = import_sessions.get(import_id)
    if not session:
        return jsonify({"error": "Import session not found"}), 404

    if session["user_id"] != user_id:
        return jsonify({"error": "Forbidden"}), 403

    # Cleanup
    if os.path.exists(session["file_path"]):
        os.remove(session["file_path"])
    del import_sessions[import_id]

    return jsonify({"message": "Import cancelled"})


@imports_bp.route("/<import_id>/ai-analyze", methods=["POST"])
@jwt_required()
def ai_analyze_structure(import_id):
    """Use AI to analyze the import structure and suggest optimal mappings."""
    user_id = get_jwt_identity()

    session = import_sessions.get(import_id)
    if not session:
        return jsonify({"error": "Import session not found"}), 404

    if session["user_id"] != user_id:
        return jsonify({"error": "Forbidden"}), 403

    ai_service = get_ai_service()
    if not ai_service.is_enabled():
        return jsonify({
            "error": "AI service not available",
            "fallback": session.get("smart_mapping", {})
        }), 200

    # Get AI analysis
    headers = session.get("headers", [])
    rows = session.get("rows", [])

    # Build raw preview from first few raw rows
    raw_preview = ""
    raw_rows = session.get("raw_rows", [])
    if raw_rows:
        for i, row in enumerate(raw_rows[:8]):
            raw_preview += f"Row {i+1}: {' | '.join(str(c)[:30] for c in row if c)}\n"

    ai_result = ai_service.analyze_spreadsheet_structure(headers, rows[:10], raw_preview)

    if ai_result:
        # Update session with AI suggestions
        session["ai_structure_analysis"] = ai_result

        # Merge AI suggestions with existing smart mapping
        ai_mapping = ai_result.get("suggested_mapping", {})
        merged_mapping = {**session.get("smart_mapping", {})}
        for key, value in ai_mapping.items():
            if value and value in headers:
                merged_mapping[key] = value

        session["ai_enhanced_mapping"] = merged_mapping

        return jsonify({
            "import_id": import_id,
            "ai_analysis": ai_result,
            "enhanced_mapping": merged_mapping,
            "original_mapping": session.get("smart_mapping", {}),
            "mapping_changed": merged_mapping != session.get("smart_mapping", {}),
        })

    return jsonify({
        "error": "AI analysis failed",
        "fallback": session.get("smart_mapping", {})
    }), 200


@imports_bp.route("/<import_id>/redetect-headers", methods=["POST"])
@jwt_required()
def redetect_headers(import_id):
    """Manually specify or re-detect the header row."""
    user_id = get_jwt_identity()

    session = import_sessions.get(import_id)
    if not session:
        return jsonify({"error": "Import session not found"}), 404

    if session["user_id"] != user_id:
        return jsonify({"error": "Forbidden"}), 403

    data = request.json or {}
    new_header_row = data.get("header_row")  # 1-indexed

    if new_header_row is None:
        return jsonify({"error": "header_row is required (1-indexed)"}), 400

    raw_rows = session.get("raw_rows", [])
    if not raw_rows:
        return jsonify({"error": "Raw data not available"}), 400

    header_row_idx = int(new_header_row) - 1  # Convert to 0-indexed

    if header_row_idx < 0 or header_row_idx >= len(raw_rows):
        return jsonify({"error": f"Invalid header row. Must be between 1 and {len(raw_rows)}"}), 400

    # Reprocess with new header row
    headers, rows = process_raw_to_structured(raw_rows, header_row_idx)

    if not headers:
        return jsonify({"error": "Could not extract headers from specified row"}), 400

    # Re-run analysis
    column_analysis = analyze_column_patterns(headers, rows)
    hierarchy_info = detect_hierarchical_structure(rows, headers)
    adjacent_pairs = detect_adjacent_content_columns(headers, rows, {})

    # Detect data types
    data_types = {}
    for header in headers:
        sample_values = [row.get(header) for row in rows[:15] if row.get(header)]
        types = [detect_data_type(v) for v in sample_values]
        non_empty_types = [t for t in types if t != "empty"]
        if non_empty_types:
            data_types[header] = max(set(non_empty_types), key=non_empty_types.count)
        else:
            data_types[header] = "text"

    # Re-detect smart mapping
    smart_mapping, confidence, confidence_details = smart_detect_column_mapping(
        headers, rows, data_types, column_analysis
    )

    # Update session
    session["header_row_idx"] = header_row_idx
    session["headers"] = headers
    session["rows"] = rows
    session["smart_mapping"] = smart_mapping
    session["mapping_confidence"] = confidence
    session["confidence_details"] = confidence_details
    session["column_analysis"] = column_analysis
    session["hierarchy_info"] = hierarchy_info
    session["adjacent_pairs"] = adjacent_pairs
    session["structure"]["headers"] = headers
    session["structure"]["data_types"] = data_types
    session["structure"]["row_count"] = len(rows)
    session["structure"]["header_row_detected"] = new_header_row

    return jsonify({
        "import_id": import_id,
        "header_row": new_header_row,
        "headers": headers,
        "row_count": len(rows),
        "preview_rows": rows[:15],
        "smart_mapping": smart_mapping,
        "mapping_confidence": confidence,
        "column_analysis": column_analysis,
    })
